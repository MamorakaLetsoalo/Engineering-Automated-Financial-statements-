# Databricks notebook source
# MAGIC %md
# MAGIC # 📤 Export Layer
# MAGIC Reads Gold Delta tables → writes CSV exports + formatted Excel workbook.
# MAGIC
# MAGIC Outputs (all to /FileStore/finmodel_pro/exports/):
# MAGIC   - historical_annual.csv
# MAGIC   - annual_forecast.csv
# MAGIC   - monthly_forecast.csv
# MAGIC   - scenarios.csv
# MAGIC   - income_statement.csv
# MAGIC   - balance_sheet.csv
# MAGIC   - cash_flow.csv
# MAGIC   - financial_model.xlsx  (4-tab formatted Excel)

# COMMAND ----------
%pip install openpyxl --quiet

# COMMAND ----------
import os, sys
import pandas as pd
import numpy as np
from datetime import datetime
from pyspark.sql import SparkSession
import pyspark.sql.functions as F

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment,
                               Border, Side, numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

spark = SparkSession.builder.getOrCreate()

try:
    sys.path.insert(0, "../configs")
    from project_config import PATHS, SCENARIOS, get_run_id
except ImportError:
    PATHS = {"gold": "/FileStore/finmodel_pro/gold",
             "silver": "/FileStore/finmodel_pro/silver",
             "exports": "/FileStore/finmodel_pro/exports"}
    def get_run_id(): return datetime.utcnow().strftime("run_%Y%m%d_%H%M%S")

try:    RUN_ID = dbutils.widgets.get("run_id")
except: RUN_ID = get_run_id()

EXPORT_LOCAL = f"/dbfs{PATHS['exports']}"
os.makedirs(EXPORT_LOCAL, exist_ok=True)
print(f"Export path: {PATHS['exports']}")

# COMMAND ----------
# MAGIC %md ## 1. Load Gold tables

# COMMAND ----------
def load_gold(name):
    try:
        return (spark.read.format("delta")
                     .load(f"{PATHS['gold']}/{name}")
                     .toPandas())
    except Exception as e:
        print(f"⚠️  Could not load gold/{name}: {e}")
        return None

def load_silver(name):
    try:
        return (spark.read.format("delta")
                     .load(f"{PATHS['silver']}/{name}")
                     .toPandas())
    except Exception as e:
        print(f"⚠️  Could not load silver/{name}: {e}")
        return None

hist      = load_silver("annual_financials")
monthly_f = load_gold("monthly_forecast")
annual_f  = load_gold("annual_forecast")
scenarios = load_gold("scenario_forecasts")
inc_stmt  = load_gold("statements/income_statement")
bal_sheet = load_gold("statements/balance_sheet")
cash_flow = load_gold("statements/cash_flow")

tables_loaded = sum(1 for t in [hist,monthly_f,annual_f,scenarios,inc_stmt,bal_sheet,cash_flow]
                    if t is not None)
print(f"✅ Loaded {tables_loaded}/7 tables from Delta")

# COMMAND ----------
# MAGIC %md ## 2. CSV Exports

# COMMAND ----------
csv_exports = {
    "historical_annual" : hist,
    "monthly_forecast"  : monthly_f,
    "annual_forecast"   : annual_f,
    "scenarios"         : scenarios,
    "income_statement"  : inc_stmt,
    "balance_sheet"     : bal_sheet,
    "cash_flow"         : cash_flow,
}

for name, df in csv_exports.items():
    if df is not None:
        path = f"{EXPORT_LOCAL}/{name}.csv"
        df.to_csv(path, index=False)
        print(f"✅ {name}.csv ({len(df)} rows)")
    else:
        print(f"⚠️  Skipped {name}.csv — no data")

# COMMAND ----------
# MAGIC %md ## 3. Excel Export — Formatted 4-tab model

# COMMAND ----------
class ExcelExporter:
    """
    Generates a professional, formatted Excel financial model.
    Tabs: Overview | Income Statement | Balance Sheet | Cash Flow | Scenarios
    """

    # ── Colour palette ─────────────────────────────────────────────────────
    DARK_BLUE  = "1F3864"
    MED_BLUE   = "2F5496"
    LIGHT_BLUE = "D9E1F2"
    FORECAST_Y = "FFF2CC"
    TOTAL_BG   = "E2EFDA"
    WHITE      = "FFFFFF"
    LIGHT_GRAY = "F2F2F2"
    GREEN      = "375623"
    RED_FONT   = "C00000"

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)   # remove default sheet

    # ── Style helpers ───────────────────────────────────────────────────────
    def _hdr(self, bg=None, fg="FFFFFF", bold=True, size=11, center=False):
        f = Font(bold=bold, color=fg, size=size)
        fill = PatternFill("solid", fgColor=bg) if bg else None
        align = Alignment(horizontal="center" if center else "left",
                          vertical="center", wrap_text=True)
        return f, fill, align

    def _border(self, style="thin", color="CCCCCC"):
        s = Side(style=style, color=color)
        return Border(bottom=s, top=s)

    def _apply(self, cell, font=None, fill=None, align=None, border=None, fmt=None):
        if font:   cell.font      = font
        if fill:   cell.fill      = fill
        if align:  cell.alignment = align
        if border: cell.border    = border
        if fmt:    cell.number_format = fmt

    def _money(self, v):
        if v is None or (isinstance(v, float) and np.isnan(v)): return None
        return round(float(v), 0)

    def _pct(self, v):
        if v is None or (isinstance(v, float) and np.isnan(v)): return None
        return round(float(v), 4)

    # ── Sheet: Cover ────────────────────────────────────────────────────────
    def add_cover(self, company="AcmeCorp", generated=None):
        ws = self.wb.create_sheet("📋 Cover")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 60

        generated = generated or datetime.now().strftime("%d %B %Y")
        ws.row_dimensions[1].height = 20

        data = [
            (3,  "FinModel Pro",                        14, self.DARK_BLUE, True),
            (4,  "Automated Financial Modelling Platform", 11, "444444", False),
            (6,  f"Company: {company}",                 11, "222222", True),
            (7,  f"Generated: {generated}",             10, "666666", False),
            (8,  "Currency: ZAR (South African Rand)",  10, "666666", False),
            (9,  "Fiscal Year End: December",           10, "666666", False),
            (11, "CONTENTS",                            11, self.DARK_BLUE, True),
            (12, "1. Income Statement (Historical + 3Y Forecast)", 10, "333333", False),
            (13, "2. Balance Sheet",                    10, "333333", False),
            (14, "3. Cash Flow Statement",              10, "333333", False),
            (15, "4. Scenario Analysis (Base / Bull / Bear)", 10, "333333", False),
            (17, "IMPORTANT NOTICE",                    10, self.RED_FONT, True),
            (18, "This model is for illustrative purposes. Not investment advice.", 9, "666666", False),
        ]
        for row, text, size, color, bold in data:
            c = ws.cell(row=row, column=1, value=text)
            c.font = Font(size=size, color=color, bold=bold)

        # Thick top border accent
        for col in range(1, 3):
            ws.cell(row=2, column=col).fill = PatternFill("solid", fgColor=self.DARK_BLUE)
        ws.row_dimensions[2].height = 5

    # ── Generic financial sheet writer ──────────────────────────────────────
    def _write_fin_sheet(self, ws, title: str, years: list,
                          hist_count: int, row_defs: list):
        """
        row_defs: list of dicts:
          { label, values, is_total, is_section, is_pct, is_spacer }
        """
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "B5"

        # Column widths
        ws.column_dimensions["A"].width = 36
        for i in range(len(years)):
            ws.column_dimensions[get_column_letter(i+2)].width = 16

        # Title row
        tc = ws.cell(row=1, column=1, value=title)
        tc.font = Font(bold=True, size=13, color=self.DARK_BLUE)

        # Subtitle
        ws.cell(row=2, column=1, value=f"AcmeCorp | ZAR | FY {years[0]}–{years[-1]}")
        ws.cell(row=2, column=1).font = Font(size=9, color="888888")

        # Year headers (row 4)
        lc, lf, la = self._hdr(bg=self.DARK_BLUE, center=True)
        ws.cell(row=4, column=1, value="").fill = PatternFill("solid", fgColor=self.DARK_BLUE)
        for i, yr in enumerate(years):
            c = ws.cell(row=4, column=i+2, value=int(yr))
            c.font = lc
            is_fc = i >= hist_count
            c.fill = PatternFill("solid", fgColor=self.MED_BLUE if is_fc else self.DARK_BLUE)
            c.alignment = Alignment(horizontal="center", vertical="center")
            if is_fc:
                # Add "F" superscript indicator
                ws.cell(row=3, column=i+2, value="▶ Forecast").font = \
                    Font(size=8, color=self.MED_BLUE, bold=True)
                ws.cell(row=3, column=i+2).alignment = Alignment(horizontal="center")

        # Data rows
        for r_offset, rdef in enumerate(row_defs):
            row = r_offset + 5

            if rdef.get("is_spacer"):
                ws.row_dimensions[row].height = 6
                continue

            label = rdef["label"]
            vals  = rdef.get("values", [])
            is_total   = rdef.get("is_total", False)
            is_section = rdef.get("is_section", False)
            is_pct     = rdef.get("is_pct", False)

            # Row height
            ws.row_dimensions[row].height = 18 if is_section else 16

            # Label cell
            lc = ws.cell(row=row, column=1, value=label)

            if is_section:
                lc.font  = Font(bold=True, size=10, color=self.WHITE)
                lc.fill  = PatternFill("solid", fgColor=self.MED_BLUE)
                for ci in range(2, len(years)+2):
                    ws.cell(row=row, column=ci).fill = \
                        PatternFill("solid", fgColor=self.MED_BLUE)
            elif is_total:
                lc.font  = Font(bold=True, size=10)
                lc.fill  = PatternFill("solid", fgColor=self.TOTAL_BG)
                lc.border = self._border("medium", "375623")
            else:
                lc.font  = Font(size=10)
                lc.fill  = PatternFill("solid", fgColor=(
                    self.LIGHT_GRAY if r_offset % 2 == 0 else self.WHITE))
            lc.alignment = Alignment(horizontal="left", vertical="center",
                                      indent=0 if (is_total or is_section) else 1)

            # Value cells
            for ci, val in enumerate(vals):
                c = ws.cell(row=row, column=ci+2, value=val)
                is_fc_col = ci >= hist_count

                if is_section:
                    c.fill = PatternFill("solid", fgColor=self.MED_BLUE)
                elif is_total:
                    c.font   = Font(bold=True, size=10)
                    c.fill   = PatternFill("solid", fgColor=self.TOTAL_BG)
                    c.border = self._border("medium", "375623")
                else:
                    c.fill = PatternFill("solid", fgColor=(
                        self.FORECAST_Y if is_fc_col else
                        (self.LIGHT_GRAY if r_offset % 2 == 0 else self.WHITE)))
                    c.font = Font(size=10, color="666666" if is_fc_col else "000000")

                c.alignment = Alignment(horizontal="right", vertical="center")
                if val is not None:
                    c.number_format = '0.0%' if is_pct else '#,##0'

    # ── Tab 1: Income Statement ─────────────────────────────────────────────
    def add_income_statement(self, hist_df, fcast_df):
        ws = self.wb.create_sheet("Income Statement")
        if hist_df is None and fcast_df is None: return

        h = hist_df.sort_values("year") if hist_df is not None else pd.DataFrame()
        f = fcast_df.sort_values("year") if fcast_df is not None else pd.DataFrame()

        years = list(h.get("year", pd.Series([]))) + list(f.get("year", pd.Series([])))
        hn    = len(h)

        def gv(df, col, default=None):
            return list(df[col]) if col in df.columns else [default]*len(df)

        h_rev = gv(h, "revenue");    f_rev = gv(f, "revenue")
        h_cogs= gv(h, "cogs");       f_cogs= gv(f, "cogs")
        h_gp  = gv(h, "gross_profit"); f_gp = gv(f, "gross_profit")
        h_opex= gv(h, "opex");       f_opex= gv(f, "opex")
        h_rd  = gv(h, "rd_expense"); f_rd  = gv(f, "rd_expense")
        h_sga = gv(h, "sga_expense");f_sga = gv(f, "sga_expense")
        h_eb  = gv(h, "ebitda");     f_eb  = gv(f, "ebitda")
        h_dep = gv(h, "depreciation");f_dep= gv(f, "depreciation")
        h_ebit= gv(h, "ebit");       f_ebit= gv(f, "ebit")
        h_int = gv(h, "interest_expense"); f_int = gv(f, "interest_expense")
        h_ebt = gv(h, "ebt");        f_ebt = gv(f, "ebt")
        h_tax = gv(h, "tax_expense");f_tax = gv(f, "tax_expense")
        h_ni  = gv(h, "net_income"); f_ni  = gv(f, "net_income")

        gm_h = [gv(h,"gross_margin_avg")[i] if "gross_margin_avg" in h.columns
                else (h_gp[i]/h_rev[i] if h_rev[i] else None) for i in range(len(h))]
        gm_f = [gv(f,"gross_margin")[i] if "gross_margin" in f.columns
                else (f_gp[i]/f_rev[i] if f_rev[i] else None) for i in range(len(f))]
        em_h = [gv(h,"ebitda_margin_avg")[i] if "ebitda_margin_avg" in h.columns
                else (h_eb[i]/h_rev[i] if h_rev[i] else None) for i in range(len(h))]
        em_f = [gv(f,"ebitda_margin")[i] if "ebitda_margin" in f.columns
                else (f_eb[i]/f_rev[i] if f_rev[i] else None) for i in range(len(f))]
        nm_h = [gv(h,"net_margin_avg")[i] if "net_margin_avg" in h.columns
                else (h_ni[i]/h_rev[i] if h_rev[i] else None) for i in range(len(h))]
        nm_f = [gv(f,"net_margin")[i] if "net_margin" in f.columns
                else (f_ni[i]/f_rev[i] if f_rev[i] else None) for i in range(len(f))]

        M = lambda a,b: [self._money(v) for v in a+b]
        P = lambda a,b: [self._pct(v)   for v in a+b]

        rows = [
            {"label":"REVENUE",                "is_section":True,  "values":[""]*len(years)},
            {"label":"Revenue",                "values":M(h_rev, f_rev)},
            {"label":"Cost of Goods Sold",     "values":M(h_cogs,f_cogs)},
            {"label":"Gross Profit",           "values":M(h_gp,  f_gp),   "is_total":True},
            {"label":"Gross Margin %",         "values":P(gm_h,  gm_f),   "is_pct":True},
            {"is_spacer":True},
            {"label":"OPERATING EXPENSES",     "is_section":True,  "values":[""]*len(years)},
            {"label":"Operating Expenses",     "values":M(h_opex,f_opex)},
            {"label":"R&D Expense",            "values":M(h_rd,  f_rd)},
            {"label":"SG&A Expense",           "values":M(h_sga, f_sga)},
            {"label":"EBITDA",                 "values":M(h_eb,  f_eb),   "is_total":True},
            {"label":"EBITDA Margin %",        "values":P(em_h,  em_f),   "is_pct":True},
            {"is_spacer":True},
            {"label":"BELOW EBITDA",           "is_section":True,  "values":[""]*len(years)},
            {"label":"Depreciation & Amort.",  "values":M(h_dep, f_dep)},
            {"label":"EBIT",                   "values":M(h_ebit,f_ebit), "is_total":True},
            {"label":"Interest Expense",       "values":M(h_int, f_int)},
            {"label":"Earnings Before Tax",    "values":M(h_ebt, f_ebt)},
            {"label":"Tax Expense (28%)",      "values":M(h_tax, f_tax)},
            {"label":"Net Income",             "values":M(h_ni,  f_ni),   "is_total":True},
            {"label":"Net Margin %",           "values":P(nm_h,  nm_f),   "is_pct":True},
        ]
        self._write_fin_sheet(ws, "Income Statement", years, hn, rows)

    # ── Tab 2: Balance Sheet ────────────────────────────────────────────────
    def add_balance_sheet(self, hist_df, fcast_df):
        ws = self.wb.create_sheet("Balance Sheet")
        if hist_df is None and fcast_df is None: return

        h = hist_df.sort_values("year") if hist_df is not None else pd.DataFrame()
        f = fcast_df.sort_values("year") if fcast_df is not None else pd.DataFrame()
        years = list(h.get("year", pd.Series([]))) + list(f.get("year", pd.Series([])))
        hn = len(h)

        def gv(df, col):
            return list(df[col]) if col in df.columns else [None]*len(df)

        M = lambda a,b: [self._money(v) for v in a+b]

        rows = [
            {"label":"ASSETS",               "is_section":True, "values":[""]*len(years)},
            {"label":"Cash & Equivalents",    "values":M(gv(h,"cash"),gv(f,"cash"))},
            {"label":"Accounts Receivable",   "values":M(gv(h,"accounts_receivable"),gv(f,"accounts_receivable"))},
            {"label":"Inventory",             "values":M(gv(h,"inventory"),gv(f,"inventory"))},
            {"label":"Total Current Assets",  "values":M(gv(h,"current_assets"),gv(f,"current_assets")), "is_total":True},
            {"label":"Net PP&E",              "values":M(gv(h,"net_ppe"),gv(f,"net_ppe"))},
            {"label":"Total Assets",          "values":M(gv(h,"total_assets"),gv(f,"total_assets")), "is_total":True},
            {"is_spacer":True},
            {"label":"LIABILITIES",           "is_section":True, "values":[""]*len(years)},
            {"label":"Accounts Payable",      "values":M(gv(h,"accounts_payable"),gv(f,"accounts_payable"))},
            {"label":"Long-term Debt",        "values":M(gv(h,"long_term_debt"),gv(f,"long_term_debt"))},
            {"label":"Total Liabilities",     "values":M(gv(h,"total_liabilities"),gv(f,"total_liabilities")), "is_total":True},
            {"is_spacer":True},
            {"label":"EQUITY",                "is_section":True, "values":[""]*len(years)},
            {"label":"Paid-in Capital",       "values":M(gv(h,"paid_in_capital"),gv(f,"paid_in_capital"))},
            {"label":"Retained Earnings",     "values":M(gv(h,"retained_earnings"),gv(f,"retained_earnings"))},
            {"label":"Total Equity",          "values":M(gv(h,"total_equity"),gv(f,"total_equity")), "is_total":True},
            {"is_spacer":True},
            {"label":"Total L + E (check)",   "values":M(gv(h,"total_assets"),gv(f,"total_assets")), "is_total":True},
        ]
        self._write_fin_sheet(ws, "Balance Sheet", years, hn, rows)

    # ── Tab 3: Cash Flow ────────────────────────────────────────────────────
    def add_cash_flow(self, hist_df, fcast_df):
        ws = self.wb.create_sheet("Cash Flow Statement")
        if hist_df is None and fcast_df is None: return

        h = hist_df.sort_values("year") if hist_df is not None else pd.DataFrame()
        f = fcast_df.sort_values("year") if fcast_df is not None else pd.DataFrame()
        years = list(h.get("year", pd.Series([]))) + list(f.get("year", pd.Series([])))
        hn = len(h)

        def gv(df, col):
            return list(df[col]) if col in df.columns else [None]*len(df)

        M = lambda a,b: [self._money(v) for v in a+b]

        rows = [
            {"label":"OPERATING ACTIVITIES",            "is_section":True,"values":[""]*len(years)},
            {"label":"Net Income",                      "values":M(gv(h,"net_income"),gv(f,"net_income"))},
            {"label":"Add: Depreciation & Amort.",      "values":M(gv(h,"depreciation"),gv(f,"depreciation"))},
            {"label":"Change in Accounts Receivable",   "values":M(gv(h,"cfo_d_ar"),gv(f,"cfo_d_ar"))},
            {"label":"Change in Inventory",             "values":M(gv(h,"cfo_d_inv"),gv(f,"cfo_d_inv"))},
            {"label":"Change in Accounts Payable",      "values":M(gv(h,"cfo_d_ap"),gv(f,"cfo_d_ap"))},
            {"label":"Cash from Operations (CFO)",      "values":M(gv(h,"cfo"),gv(f,"cfo")), "is_total":True},
            {"is_spacer":True},
            {"label":"INVESTING ACTIVITIES",            "is_section":True,"values":[""]*len(years)},
            {"label":"Capital Expenditures",            "values":M(gv(h,"capex_outflow"),gv(f,"cfi"))},
            {"label":"Cash from Investing (CFI)",       "values":M(gv(h,"cfi"),gv(f,"cfi")), "is_total":True},
            {"is_spacer":True},
            {"label":"FINANCING ACTIVITIES",            "is_section":True,"values":[""]*len(years)},
            {"label":"Dividends Paid",                  "values":M(gv(h,"dividends"),gv(f,"dividends"))},
            {"label":"Net Debt Change",                 "values":M(gv(h,"debt_change"),gv(f,"debt_change"))},
            {"label":"Cash from Financing (CFF)",       "values":M(gv(h,"cff"),gv(f,"cff")), "is_total":True},
            {"is_spacer":True},
            {"label":"Net Change in Cash",              "values":M(gv(h,"net_cash_change"),gv(f,"net_cash_change")), "is_total":True},
            {"label":"Closing Cash Balance",            "values":M(gv(h,"closing_cash"),gv(f,"closing_cash")), "is_total":True},
        ]
        self._write_fin_sheet(ws, "Cash Flow Statement (Indirect Method)", years, hn, rows)

    # ── Tab 4: Scenarios ────────────────────────────────────────────────────
    def add_scenarios(self, scenarios_df, fcast_years):
        ws = self.wb.create_sheet("Scenario Analysis")
        if scenarios_df is None: return

        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 28

        sc_labels = ["Base", "Bull", "Bear"]
        sc_colors = {
            "Base": ("2F5496", "D9E1F2"),
            "Bull": ("375623", "E2EFDA"),
            "Bear": ("843C0C", "FCE4D6"),
        }

        # Header
        tc = ws.cell(row=1, column=1, value="Scenario Analysis — Base / Bull / Bear")
        tc.font = Font(bold=True, size=13, color=self.DARK_BLUE)

        col = 2
        for sc in sc_labels:
            dark, light = sc_colors[sc]
            for yi, yr in enumerate(sorted(fcast_years)):
                c = ws.cell(row=3, column=col, value=int(yr))
                c.font  = Font(bold=True, color="FFFFFF", size=10)
                c.fill  = PatternFill("solid", fgColor=dark)
                c.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(col)].width = 16
                col += 1

        # Scenario label spans
        col = 2
        for sc in sc_labels:
            dark, light = sc_colors[sc]
            c = ws.cell(row=2, column=col, value=f"◆ {sc} Case")
            c.font  = Font(bold=True, color="FFFFFF", size=11)
            c.fill  = PatternFill("solid", fgColor=dark)
            c.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=2, start_column=col,
                           end_row=2,   end_column=col+len(fcast_years)-1)
            col += len(fcast_years)

        # Metrics
        metrics = [
            ("Revenue",        "revenue",       False),
            ("Gross Profit",   "gross_profit",  False),
            ("EBITDA",         "ebitda",        False),
            ("EBITDA Margin",  "ebitda_margin", True),
            ("Net Income",     "net_income",    False),
            ("Net Margin",     "net_margin",    True),
            ("Free Cash Flow", "fcf",           False),
        ]

        for r_offset, (label, col_name, is_pct) in enumerate(metrics):
            row = r_offset + 4
            lc = ws.cell(row=row, column=1, value=label)
            lc.font  = Font(bold=False, size=10)
            lc.fill  = PatternFill("solid", fgColor=(self.LIGHT_GRAY if r_offset%2==0 else self.WHITE))
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)

            col = 2
            for sc in sc_labels:
                dark, light = sc_colors[sc]
                sub = scenarios_df[scenarios_df["scenario"]==sc].sort_values("year")
                for yr in sorted(fcast_years):
                    row_data = sub[sub["year"]==yr]
                    val = row_data[col_name].values[0] if (len(row_data)>0 and col_name in sub.columns) else None
                    c = ws.cell(row=row, column=col,
                                value=self._pct(val) if is_pct else self._money(val))
                    c.fill = PatternFill("solid", fgColor=light)
                    c.alignment = Alignment(horizontal="right")
                    c.font = Font(size=10)
                    c.number_format = '0.0%' if is_pct else '#,##0'
                    col += 1

    def save(self, path: str):
        self.wb.save(path)
        print(f"✅ Excel saved: {path}")


# COMMAND ----------
# MAGIC %md ## 4. Generate Excel workbook

# COMMAND ----------
exporter = ExcelExporter()
exporter.add_cover(company="AcmeCorp")

# Income statement: use gold statements if available, else silver+gold forecast
if inc_stmt is not None:
    exporter.add_income_statement(
        hist_df  = inc_stmt[inc_stmt.get("period_type","Forecast") == "Historical"] if "period_type" in inc_stmt.columns else inc_stmt,
        fcast_df = inc_stmt[inc_stmt.get("period_type","Historical") == "Forecast"] if "period_type" in inc_stmt.columns else annual_f,
    )
else:
    exporter.add_income_statement(hist_df=hist, fcast_df=annual_f)

exporter.add_balance_sheet(hist_df=hist, fcast_df=annual_f)
exporter.add_cash_flow(hist_df=hist, fcast_df=annual_f)

if scenarios is not None:
    fcast_years = sorted(scenarios["year"].unique()) if "year" in scenarios.columns else [2024,2025,2026]
    exporter.add_scenarios(scenarios, fcast_years)

xlsx_path = f"{EXPORT_LOCAL}/financial_model.xlsx"
exporter.save(xlsx_path)

# COMMAND ----------
# MAGIC %md ## 5. Verify exports

# COMMAND ----------
print("\n" + "="*55)
print("EXPORT SUMMARY")
print("="*55)
for f in os.listdir(EXPORT_LOCAL):
    full = os.path.join(EXPORT_LOCAL, f)
    size = os.path.getsize(full) / 1024
    print(f"  ✅ {f:<35} {size:>8.1f} KB")

print(f"\n📁 All exports at: {PATHS['exports']}")
print("\nTo use in Streamlit: copy CSV files to ./exports/")
print("To use in Power BI:  import CSV files via Get Data → Text/CSV")

try:
    dbutils.jobs.taskValues.set("export_status", "success")
    dbutils.jobs.taskValues.set("xlsx_path", xlsx_path)
except: pass
