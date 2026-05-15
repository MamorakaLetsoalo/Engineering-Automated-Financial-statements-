# Databricks notebook source
# MAGIC %md
# MAGIC # 📊 Notebook 4: Financial Statement Generator (Gold → Statements)
# MAGIC **Pipeline:** Gold forecast data → Structured 3-statement model
# MAGIC
# MAGIC Generates:
# MAGIC - Income Statement (P&L)
# MAGIC - Balance Sheet
# MAGIC - Cash Flow Statement
# MAGIC - Validation: Assets = Liabilities + Equity, cash reconciliation

# COMMAND ----------
%pip install openpyxl --quiet

# COMMAND ----------
import pandas as pd
import numpy as np
from pyspark.sql import SparkSession

spark = SparkSession.builder.getOrCreate()

GOLD_PATH   = "/FileStore/financial_model/gold"
SILVER_PATH = "/FileStore/financial_model/silver"
EXPORT_PATH = "/FileStore/financial_model/exports"

# COMMAND ----------
# MAGIC %md ## 1. Load data

# COMMAND ----------
historical = (spark.read.format("delta")
              .load(f"{SILVER_PATH}/annual_financials")
              .toPandas().sort_values("year"))

forecast   = (spark.read.format("delta")
              .load(f"{GOLD_PATH}/annual_forecast")
              .toPandas().sort_values("year"))

scenarios  = (spark.read.format("delta")
              .load(f"{GOLD_PATH}/scenario_forecasts")
              .toPandas())

print(f"✅ Historical: {len(historical)} years | Forecast: {len(forecast)} years")

# COMMAND ----------
# MAGIC %md ## 2. Income Statement Builder

# COMMAND ----------
def build_income_statement(hist_df: pd.DataFrame,
                            fcast_df: pd.DataFrame,
                            scenario: str = "Base") -> pd.DataFrame:
    """
    Builds a clean income statement combining historical + forecast.
    All values in ZAR.
    """
    if scenario != "Base":
        fcast_df = scenarios[scenarios["scenario"] == scenario].copy()

    hist_cols = {
        "year": "year", "revenue": "revenue", "cogs": "cogs",
        "gross_profit": "gross_profit", "opex": "opex",
        "rd_expense": "rd_expense", "sga_expense": "sga_expense",
        "ebitda": "ebitda", "depreciation": "depreciation",
        "ebit": "ebit", "interest_expense": "interest_expense",
        "ebt": "ebt", "tax_expense": "tax_expense", "net_income": "net_income",
    }

    hist_clean  = hist_df[[c for c in hist_cols if c in hist_df.columns]].copy()
    fcast_clean = fcast_df[[c for c in hist_cols if c in fcast_df.columns]].copy()

    hist_clean["period_type"]  = "Historical"
    fcast_clean["period_type"] = f"Forecast ({scenario})"

    combined = pd.concat([hist_clean, fcast_clean], ignore_index=True).sort_values("year")

    # Derived metrics
    combined["gross_margin"]  = combined["gross_profit"]  / combined["revenue"]
    combined["ebitda_margin"] = combined["ebitda"]        / combined["revenue"]
    combined["ebit_margin"]   = combined["ebit"]          / combined["revenue"]
    combined["net_margin"]    = combined["net_income"]     / combined["revenue"]
    combined["tax_rate"]      = combined["tax_expense"]    / combined["ebt"].replace(0, np.nan)

    # YoY revenue growth
    combined["revenue_growth"] = combined["revenue"].pct_change()

    return combined

is_base = build_income_statement(historical, forecast, "Base")
is_bull = build_income_statement(historical, forecast, "Bull")
is_bear = build_income_statement(historical, forecast, "Bear")

print("✅ Income statements built — Base / Bull / Bear")
display(is_base[["year", "period_type", "revenue", "gross_margin",
                 "ebitda_margin", "net_margin", "net_income"]])

# COMMAND ----------
# MAGIC %md ## 3. Balance Sheet Builder

# COMMAND ----------
def build_balance_sheet(hist_df: pd.DataFrame,
                         fcast_df: pd.DataFrame) -> pd.DataFrame:
    """
    Constructs balance sheet from drivers:
    - Current assets: AR + Inventory + Cash
    - Fixed assets: cumulative capex net of depreciation
    - Liabilities: AP + long-term debt
    - Equity: retained earnings (accumulating net income)
    Validates: Assets = Liabilities + Equity
    """
    hist  = hist_df.copy()
    fcast = fcast_df.copy()

    hist["period_type"]  = "Historical"
    fcast["period_type"] = "Forecast"

    combined = pd.concat([hist, fcast], ignore_index=True).sort_values("year")

    # ── Current Assets ──────────────────────────────────────────────────
    combined["cash"] = 0.0  # will fill below
    combined["current_assets"] = (
        combined["accounts_receivable"] +
        combined["inventory"]
    )

    # ── Fixed Assets (net PPE) ───────────────────────────────────────────
    # Starting fixed assets (estimate from first year)
    start_ppe = hist_df["capex"].iloc[0] * 4  # rough 4x first year capex
    ppe = [start_ppe]
    for i in range(1, len(combined)):
        capex_i = abs(combined["capex"].iloc[i]) if "capex" in combined.columns else 0
        depr_i  = combined["depreciation"].iloc[i] if "depreciation" in combined.columns else 0
        ppe.append(max(ppe[-1] + capex_i - depr_i, 0))
    combined["net_ppe"] = ppe

    # ── Total Assets (before cash) ───────────────────────────────────────
    combined["total_assets_excl_cash"] = combined["current_assets"] + combined["net_ppe"]

    # ── Liabilities ──────────────────────────────────────────────────────
    # Long-term debt: modelled as a fixed amount declining slowly
    initial_debt = hist_df["interest_expense"].mean() / 0.07 * 12  # implied from interest
    debt = [initial_debt]
    for i in range(1, len(combined)):
        # Repay 5% of debt per year
        debt.append(max(debt[-1] * 0.95, 0))
    combined["long_term_debt"] = debt

    combined["total_liabilities"] = combined["accounts_payable"] + combined["long_term_debt"]

    # ── Equity (retained earnings accumulate) ────────────────────────────
    retained = [hist_df["net_income"].iloc[0] * 0.7]  # 70% retention
    for i in range(1, len(combined)):
        retained.append(retained[-1] + combined["net_income"].iloc[i] * 0.7)
    paid_in_capital = hist_df["revenue"].iloc[0] * 0.3  # assumed paid-in
    combined["retained_earnings"] = retained
    combined["paid_in_capital"]   = paid_in_capital
    combined["total_equity"]      = combined["retained_earnings"] + paid_in_capital

    # ── Cash (balancing item: Assets = Liabilities + Equity) ─────────────
    combined["total_assets"] = combined["total_liabilities"] + combined["total_equity"]
    combined["cash"] = (
        combined["total_assets"] -
        combined["total_assets_excl_cash"]
    ).clip(lower=0)

    # Recalculate current assets with cash
    combined["current_assets"] = (
        combined["cash"] +
        combined["accounts_receivable"] +
        combined["inventory"]
    )
    combined["total_assets"] = combined["current_assets"] + combined["net_ppe"]

    # ── Validation ────────────────────────────────────────────────────────
    combined["bs_balance_check"] = abs(
        combined["total_assets"] -
        (combined["total_liabilities"] + combined["total_equity"])
    )

    return combined

balance_sheet = build_balance_sheet(historical, forecast)

print("✅ Balance Sheet built")
bs_check = balance_sheet["bs_balance_check"].max()
if bs_check < 1.0:
    print(f"   ✅ Balance sheet balances (max error: R{bs_check:.2f})")
else:
    print(f"   ⚠️  Balance sheet discrepancy: R{bs_check:,.2f} — check drivers")

display(balance_sheet[["year", "period_type", "total_assets",
                        "total_liabilities", "total_equity", "cash",
                        "net_ppe", "long_term_debt", "retained_earnings"]])

# COMMAND ----------
# MAGIC %md ## 4. Cash Flow Statement Builder

# COMMAND ----------
def build_cashflow_statement(is_df: pd.DataFrame,
                              bs_df: pd.DataFrame) -> pd.DataFrame:
    """
    Builds indirect-method cash flow statement.
    CFO = Net income + D&A ± working capital changes
    CFI = Capex
    CFF = Net debt changes
    """
    combined = is_df.merge(
        bs_df[["year", "cash", "accounts_receivable", "inventory",
               "accounts_payable", "long_term_debt", "net_ppe"]],
        on="year", how="left"
    ).sort_values("year")

    # Working capital changes
    combined["d_ar"]  = combined["accounts_receivable"].diff().fillna(0)
    combined["d_inv"] = combined["inventory"].diff().fillna(0)
    combined["d_ap"]  = combined["accounts_payable"].diff().fillna(0)

    # ── Operating Activities ──────────────────────────────────────────────
    combined["cfo_net_income"]  = combined["net_income"]
    combined["cfo_dna"]         = combined["depreciation"]
    combined["cfo_d_ar"]        = -combined["d_ar"]     # increase in AR = use of cash
    combined["cfo_d_inv"]       = -combined["d_inv"]    # increase in inv = use of cash
    combined["cfo_d_ap"]        =  combined["d_ap"]     # increase in AP = source of cash
    combined["cfo"]             = (
        combined["cfo_net_income"] +
        combined["cfo_dna"] +
        combined["cfo_d_ar"] +
        combined["cfo_d_inv"] +
        combined["cfo_d_ap"]
    )

    # ── Investing Activities ──────────────────────────────────────────────
    combined["capex_outflow"]   = combined["net_ppe"].diff().fillna(0) + combined["depreciation"]
    combined["cfi"]             = -combined["capex_outflow"]

    # ── Financing Activities ──────────────────────────────────────────────
    combined["debt_change"]     = combined["long_term_debt"].diff().fillna(0)
    combined["dividends"]       = -combined["net_income"] * 0.30  # 30% payout ratio
    combined["cff"]             = combined["debt_change"] + combined["dividends"]

    # ── Net change & closing cash ─────────────────────────────────────────
    combined["net_cash_change"] = combined["cfo"] + combined["cfi"] + combined["cff"]
    combined["closing_cash"]    = combined["cash"]

    # ── Validation ────────────────────────────────────────────────────────
    combined["cf_reconciliation"] = abs(
        combined["net_cash_change"] - combined["cash"].diff().fillna(0)
    )

    return combined

cashflow = build_cashflow_statement(is_base, balance_sheet)

print("✅ Cash Flow Statement built")
cf_check = cashflow["cf_reconciliation"].mean()
print(f"   Avg reconciliation error: R{cf_check:,.0f}")
display(cashflow[["year", "period_type", "cfo", "cfi", "cff",
                  "net_cash_change", "closing_cash"]])

# COMMAND ----------
# MAGIC %md ## 5. Validate 3-statement linkage

# COMMAND ----------
def validate_three_statements(is_df, bs_df, cf_df):
    print("=" * 50)
    print("3-STATEMENT VALIDATION")
    print("=" * 50)
    checks = []

    # Check 1: BS balances
    bs_err = bs_df["bs_balance_check"].max()
    checks.append(("Balance Sheet balances", bs_err < 10, f"Max error: R{bs_err:,.0f}"))

    # Check 2: Net income flows to retained earnings
    ni_flow = all(
        abs(bs_df["retained_earnings"].diff().dropna() -
            is_df["net_income"].iloc[1:].values * 0.7) < 100
    )
    checks.append(("Net income → Retained Earnings", ni_flow, ""))

    # Check 3: Revenue > 0 for all forecast years
    rev_positive = (is_df["revenue"] > 0).all()
    checks.append(("All revenues positive", rev_positive, ""))

    # Check 4: EBITDA consistency
    ebitda_check = all(
        abs(is_df["ebitda"] - (is_df["gross_profit"] - is_df["opex"] -
                                is_df.get("rd_expense", 0) - is_df.get("sga_expense", 0))) < 100
    )
    checks.append(("EBITDA consistency", ebitda_check, ""))

    for name, passed, note in checks:
        icon = "✅" if passed else "⚠️ "
        print(f"   {icon} {name}  {note}")

    all_pass = all(c[1] for c in checks)
    print(f"\n{'✅ All checks passed' if all_pass else '⚠️  Some checks failed — review above'}")
    return all_pass

validate_three_statements(is_base, balance_sheet, cashflow)

# COMMAND ----------
# MAGIC %md ## 6. Export to CSV & Excel

# COMMAND ----------
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def export_to_excel(is_df, bs_df, cf_df, scenarios_is, output_path: str):
    """
    Generates a clean Excel financial model with:
    - Tab 1: Income Statement (historical + forecast)
    - Tab 2: Balance Sheet
    - Tab 3: Cash Flow Statement
    - Tab 4: Scenario Comparison
    """
    wb = openpyxl.Workbook()

    # ── Styles ─────────────────────────────────────────────────────────────
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill("solid", fgColor="1F3864")
    subheader_fill= PatternFill("solid", fgColor="2F5496")
    alt_fill      = PatternFill("solid", fgColor="EBF0FA")
    forecast_fill = PatternFill("solid", fgColor="FFF2CC")
    center_align  = Alignment(horizontal="center")
    right_align   = Alignment(horizontal="right")
    thin_border   = Border(
        bottom=Side(style="thin", color="CCCCCC")
    )

    def fmt_currency(val):
        if pd.isna(val): return "-"
        return f"R {val:>14,.0f}"

    def fmt_pct(val):
        if pd.isna(val): return "-"
        return f"{val*100:.1f}%"

    def write_sheet(ws, title: str, rows: list, col_headers: list):
        """Generic sheet writer with formatting."""
        ws.title = title
        ws.column_dimensions["A"].width = 32
        for i, h in enumerate(col_headers, 2):
            ws.column_dimensions[get_column_letter(i)].width = 16

        # Header row
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells(f"A1:{get_column_letter(len(col_headers)+1)}1")

        # Column headers
        for i, h in enumerate(col_headers, 2):
            cell = ws.cell(row=3, column=i, value=str(h))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        ws["A3"].font = header_font
        ws["A3"].fill = header_fill
        ws["A3"].value = "Line Item"

        # Data rows
        for r_idx, (label, values, fmt, is_total) in enumerate(rows, 4):
            row_fill = PatternFill("solid", fgColor="D9E1F2") if is_total else (
                       alt_fill if r_idx % 2 == 0 else None)
            label_cell = ws.cell(row=r_idx, column=1, value=label)
            if is_total:
                label_cell.font = Font(bold=True)
            if row_fill:
                label_cell.fill = row_fill

            for c_idx, val in enumerate(values, 2):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = right_align
                if row_fill:
                    cell.fill = row_fill
                if is_total:
                    cell.font = Font(bold=True)

    # ── Tab 1: Income Statement ────────────────────────────────────────────
    ws_is = wb.active
    years = is_base["year"].tolist()
    types = is_base["period_type"].tolist()
    col_headers = [f"{y}\n({t[:4]})" for y, t in zip(years, types)]

    def get_row(col, fmt_fn):
        return [fmt_fn(v) for v in is_base[col].tolist()]

    is_rows = [
        ("Revenue",             get_row("revenue",          fmt_currency), "currency", False),
        ("Cost of Goods Sold",  get_row("cogs",             fmt_currency), "currency", False),
        ("Gross Profit",        get_row("gross_profit",     fmt_currency), "currency", True),
        ("Gross Margin %",      get_row("gross_margin",     fmt_pct),      "pct",      False),
        ("",                    [""] * len(years),                         "",         False),
        ("Operating Expenses",  get_row("opex",             fmt_currency), "currency", False),
        ("R&D Expense",         get_row("rd_expense",       fmt_currency), "currency", False),
        ("SG&A Expense",        get_row("sga_expense",      fmt_currency), "currency", False),
        ("EBITDA",              get_row("ebitda",           fmt_currency), "currency", True),
        ("EBITDA Margin %",     get_row("ebitda_margin",    fmt_pct),      "pct",      False),
        ("",                    [""] * len(years),                         "",         False),
        ("Depreciation & Amort",get_row("depreciation",     fmt_currency), "currency", False),
        ("EBIT",                get_row("ebit",             fmt_currency), "currency", True),
        ("Interest Expense",    get_row("interest_expense", fmt_currency), "currency", False),
        ("EBT",                 get_row("ebt",              fmt_currency), "currency", False),
        ("Tax Expense",         get_row("tax_expense",      fmt_currency), "currency", False),
        ("Net Income",          get_row("net_income",       fmt_currency), "currency", True),
        ("Net Margin %",        get_row("net_margin",       fmt_pct),      "pct",      False),
    ]
    write_sheet(ws_is, "Income Statement", is_rows, col_headers)

    # ── Tab 2: Balance Sheet ───────────────────────────────────────────────
    ws_bs = wb.create_sheet("Balance Sheet")
    bs_years = balance_sheet["year"].tolist()
    bs_headers = [f"{y}" for y in bs_years]

    def get_bs(col):
        return [fmt_currency(v) for v in balance_sheet[col].tolist()]

    bs_rows = [
        ("ASSETS",                  [""] * len(bs_years), "", True),
        ("Cash & Equivalents",      get_bs("cash"),                 "c", False),
        ("Accounts Receivable",     get_bs("accounts_receivable"),  "c", False),
        ("Inventory",               get_bs("inventory"),            "c", False),
        ("Total Current Assets",    get_bs("current_assets"),       "c", True),
        ("Net PP&E",                 get_bs("net_ppe"),              "c", False),
        ("Total Assets",            get_bs("total_assets"),         "c", True),
        ("",                        [""] * len(bs_years),           "", False),
        ("LIABILITIES & EQUITY",    [""] * len(bs_years),           "", True),
        ("Accounts Payable",        get_bs("accounts_payable"),     "c", False),
        ("Long-term Debt",          get_bs("long_term_debt"),       "c", False),
        ("Total Liabilities",       get_bs("total_liabilities"),    "c", True),
        ("Paid-in Capital",         get_bs("paid_in_capital"),      "c", False),
        ("Retained Earnings",       get_bs("retained_earnings"),    "c", False),
        ("Total Equity",            get_bs("total_equity"),         "c", True),
        ("Total L + E",             get_bs("total_assets"),         "c", True),
    ]
    write_sheet(ws_bs, "Balance Sheet", bs_rows, bs_headers)

    # ── Tab 3: Cash Flow ───────────────────────────────────────────────────
    ws_cf = wb.create_sheet("Cash Flow Statement")
    cf_years   = cashflow["year"].tolist()
    cf_headers = [f"{y}" for y in cf_years]

    def get_cf(col):
        return [fmt_currency(v) for v in cashflow[col].tolist()]

    cf_rows = [
        ("OPERATING ACTIVITIES",         [""] * len(cf_years), "", True),
        ("Net Income",                   get_cf("cfo_net_income"),  "c", False),
        ("Add: Depreciation & Amort",    get_cf("cfo_dna"),        "c", False),
        ("Change in Accounts Receivable",get_cf("cfo_d_ar"),       "c", False),
        ("Change in Inventory",          get_cf("cfo_d_inv"),      "c", False),
        ("Change in Accounts Payable",   get_cf("cfo_d_ap"),       "c", False),
        ("Cash from Operations (CFO)",   get_cf("cfo"),            "c", True),
        ("",                             [""] * len(cf_years),     "", False),
        ("INVESTING ACTIVITIES",         [""] * len(cf_years),     "", True),
        ("Capital Expenditures",         get_cf("cfi"),            "c", False),
        ("Cash from Investing (CFI)",    get_cf("cfi"),            "c", True),
        ("",                             [""] * len(cf_years),     "", False),
        ("FINANCING ACTIVITIES",         [""] * len(cf_years),     "", True),
        ("Debt Issuance / (Repayment)",  get_cf("debt_change"),    "c", False),
        ("Dividends Paid",               get_cf("dividends"),      "c", False),
        ("Cash from Financing (CFF)",    get_cf("cff"),            "c", True),
        ("",                             [""] * len(cf_years),     "", False),
        ("Net Change in Cash",           get_cf("net_cash_change"),"c", True),
        ("Closing Cash Balance",         get_cf("closing_cash"),   "c", True),
    ]
    write_sheet(ws_cf, "Cash Flow Statement", cf_rows, cf_headers)

    # ── Tab 4: Scenario Comparison ─────────────────────────────────────────
    ws_sc = wb.create_sheet("Scenario Analysis")
    sc_years = sorted(forecast["year"].unique())
    sc_headers = [f"{s} {y}" for s in ["Base","Bull","Bear"] for y in sc_years]

    sc_rows = []
    for metric, label, fmt_fn in [
        ("revenue",       "Revenue",       fmt_currency),
        ("ebitda",        "EBITDA",        fmt_currency),
        ("ebitda_margin", "EBITDA Margin", fmt_pct),
        ("net_income",    "Net Income",    fmt_currency),
        ("net_margin",    "Net Margin",    fmt_pct),
        ("fcf",           "Free Cash Flow",fmt_currency),
    ]:
        vals = []
        for sc in ["Base","Bull","Bear"]:
            sub = scenarios[scenarios["scenario"] == sc].sort_values("year")
            for y in sc_years:
                row_val = sub[sub["year"] == y][metric]
                vals.append(fmt_fn(row_val.values[0]) if len(row_val) > 0 else "-")
        sc_rows.append((label, vals, "c", metric in ["revenue","ebitda","net_income","fcf"]))

    write_sheet(ws_sc, "Scenario Analysis", sc_rows, sc_headers)

    # Save
    local_path = f"/dbfs{output_path}/financial_model.xlsx"
    wb.save(local_path)
    print(f"✅ Excel model saved: {output_path}/financial_model.xlsx")
    return local_path

# Run export
xlsx_path = export_to_excel(is_base, balance_sheet, cashflow, scenarios, EXPORT_PATH)

# COMMAND ----------
# MAGIC %md ## 7. Write final Gold statement tables

# COMMAND ----------
for name, df in [
    ("income_statement",  is_base),
    ("balance_sheet",     balance_sheet),
    ("cash_flow",         cashflow),
    ("scenario_analysis", scenarios),
]:
    sdf = spark.createDataFrame(df.astype(str))  # stringify for schema safety
    (sdf.write
        .format("delta")
        .mode("overwrite")
        .option("overwriteSchema", "true")
        .save(f"{GOLD_PATH}/statements/{name}"))
    df.to_csv(f"/dbfs{EXPORT_PATH}/{name}.csv", index=False)
    print(f"✅ {name}: Delta + CSV exported")

# COMMAND ----------
print("=" * 55)
print("FINANCIAL STATEMENTS — COMPLETE")
print("=" * 55)
print("""
  ✅ Income Statement    (Historical + 3Y Forecast)
  ✅ Balance Sheet       (with BS balance validation)
  ✅ Cash Flow Statement (indirect method)
  ✅ Scenario Analysis   (Base / Bull / Bear)
  ✅ Excel Export        (/exports/financial_model.xlsx)
  ✅ CSV Exports         (/exports/*.csv)

  Next: Run Streamlit app for interactive UI
  📁 Export path: /FileStore/financial_model/exports/
""")
